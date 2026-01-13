from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
from datetime import datetime, timedelta
import io
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import json
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'paystub-secret-key-2025-change-this'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# 2025 Federal Tax Brackets (Single Filer)
FEDERAL_BRACKETS_2025 = [
    (0, 11600, 0.10),
    (11600, 47150, 0.12),
    (47150, 100525, 0.22),
    (100525, 191950, 0.24),
    (191950, 243725, 0.32),
    (243725, 609350, 0.35),
    (609350, float('inf'), 0.37)
]

STANDARD_DEDUCTION_2025 = 14600

# State Tax Rates 2025
STATE_TAX_RATES = {
    'AL': 0.05, 'AK': 0.00, 'AZ': 0.0575, 'AR': 0.0675, 'CA': 0.093,
    'CO': 0.0465, 'CT': 0.0663, 'DE': 0.066, 'FL': 0.00, 'GA': 0.0575,
    'HI': 0.0815, 'ID': 0.0585, 'IL': 0.0495, 'IN': 0.0325, 'IA': 0.0605,
    'KS': 0.057, 'KY': 0.05, 'LA': 0.04, 'ME': 0.065, 'MD': 0.0575,
    'MA': 0.05, 'MI': 0.0425, 'MN': 0.0785, 'MS': 0.05, 'MO': 0.055,
    'MT': 0.0685, 'NE': 0.0684, 'NV': 0.00, 'NH': 0.00, 'NJ': 0.0637,
    'NM': 0.05, 'NY': 0.065, 'NC': 0.0525, 'ND': 0.00, 'OH': 0.0515,
    'OK': 0.055, 'OR': 0.0895, 'PA': 0.0307, 'RI': 0.0675, 'SC': 0.07,
    'SD': 0.00, 'TN': 0.00, 'TX': 0.00, 'UT': 0.0495, 'VT': 0.0875,
    'VA': 0.0575, 'WA': 0.00, 'WV': 0.065, 'WI': 0.0685, 'WY': 0.00
}

class User(UserMixin):
    def __init__(self, id, email):
        self.id = id
        self.email = email

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("SELECT id, email FROM users WHERE id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return User(row[0], row[1]) if row else None

def init_db():
    """Initialize database with all required tables"""
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Companies table
    c.execute('''CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        ein TEXT,
        address TEXT,
        phone TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    # Employees table
    c.execute('''CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        company_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        ssn TEXT,
        street TEXT,
        city TEXT,
        state TEXT,
        zip TEXT,
        pay_rate REAL NOT NULL,
        is_hourly INTEGER DEFAULT 1,
        pay_frequency TEXT DEFAULT 'biweekly',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    )''')
    
    # Paystubs table
    c.execute('''CREATE TABLE IF NOT EXISTS paystubs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        check_number INTEGER NOT NULL,
        period_start DATE NOT NULL,
        period_end DATE NOT NULL,
        hours_worked REAL DEFAULT 80,
        gross_pay REAL NOT NULL,
        federal_tax REAL NOT NULL,
        state_tax REAL NOT NULL,
        social_security REAL NOT NULL,
        medicare REAL NOT NULL,
        other_deductions REAL DEFAULT 0,
        net_pay REAL NOT NULL,
        ytd_gross REAL NOT NULL,
        ytd_federal REAL NOT NULL,
        ytd_state REAL NOT NULL,
        ytd_fica REAL NOT NULL,
        ytd_net REAL NOT NULL,
        pdf_blob BLOB,
        edited INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    # Stub edits history
    c.execute('''CREATE TABLE IF NOT EXISTS stub_edits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        paystub_id INTEGER NOT NULL,
        field_name TEXT NOT NULL,
        old_value TEXT,
        new_value TEXT,
        propagate_to_later INTEGER DEFAULT 0,
        edited_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(paystub_id) REFERENCES paystubs(id) ON DELETE CASCADE
    )''')
    
    # Transactions table
    c.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        paystub_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        transaction_date DATE NOT NULL,
        description TEXT NOT NULL,
        merchant TEXT,
        category TEXT NOT NULL,
        amount REAL NOT NULL,
        transaction_type TEXT NOT NULL,
        location_city TEXT,
        location_state TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(paystub_id) REFERENCES paystubs(id) ON DELETE CASCADE,
        FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    # Employee deductions table (401k, medical, etc.)
    c.execute('''CREATE TABLE IF NOT EXISTS employee_deductions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        deduction_name TEXT NOT NULL,
        deduction_type TEXT NOT NULL,
        amount REAL NOT NULL,
        is_percentage INTEGER DEFAULT 0,
        is_pre_tax INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    conn.commit()
    conn.close()

init_db()

def calculate_federal_tax(gross, ytd_gross=0, pay_frequency='biweekly'):
    """Calculate 2025 federal income tax with progressive brackets"""
    # Determine pay periods per year based on frequency
    periods_map = {'weekly': 52, 'biweekly': 26, 'semimonthly': 24, 'monthly': 12}
    pay_periods_per_year = periods_map.get(pay_frequency, 26)
    
    # Calculate taxable income after standard deduction
    standard_deduction_per_period = STANDARD_DEDUCTION_2025 / pay_periods_per_year
    taxable = max(0, gross - standard_deduction_per_period)
    
    if taxable <= 0:
        return 0
    
    tax = 0
    
    # Process each tax bracket (progressive tax system)
    for lower, upper, rate in FEDERAL_BRACKETS_2025:
        if taxable <= lower:
            # Income is below this bracket, we're done
            break
        
        # Calculate how much of the taxable income falls in this bracket
        income_in_bracket = min(taxable, upper) - lower
        
        if income_in_bracket > 0:
            tax += income_in_bracket * rate
    
    return max(0, round(tax, 2))

def calculate_state_tax(gross, state):
    """Calculate state income tax"""
    rate = STATE_TAX_RATES.get(state, 0)
    return gross * rate

def calculate_fica(gross, ytd_gross=0):
    """Calculate Social Security (6.2% up to $168,600 annual) and Medicare (1.45%)"""
    ss_wage_base = 168600  # 2025 Social Security wage base limit
    
    # Calculate Social Security tax
    # Only apply to income up to the annual limit, accounting for YTD
    taxable_for_ss = max(0, min(gross, ss_wage_base - ytd_gross))
    social_security = taxable_for_ss * 0.062 if taxable_for_ss > 0 else 0
    
    # Medicare tax applies to all income (no limit for regular Medicare)
    medicare = gross * 0.0145
    
    return social_security, medicare

def calculate_paystub(employee_data, hours, gross_override=None, ytd_gross=0):
    """Calculate complete paystub with all deductions and YTD"""
    employee_id = employee_data['id']
    pay_rate = employee_data['pay_rate']
    is_hourly = employee_data['is_hourly']
    state = employee_data['state']
    
    # Calculate gross pay
    if gross_override:
        gross = gross_override
    elif is_hourly:
        gross = hours * pay_rate
    else:
        # Salary: annual divided by pay periods
        periods = {'weekly': 52, 'biweekly': 26, 'semimonthly': 24, 'monthly': 12}
        gross = pay_rate / periods.get(employee_data['pay_frequency'], 26)
    
    # Get employee deductions
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("""
        SELECT deduction_name, deduction_type, amount, is_percentage, is_pre_tax
        FROM employee_deductions
        WHERE employee_id=?
    """, (employee_id,))
    deductions = c.fetchall()
    conn.close()
    
    # Calculate pre-tax deductions (401k, health insurance, etc.)
    pre_tax_deductions = {}
    pre_tax_total = 0
    ss_taxable_reduction = 0  # Only 401k reduces SS taxable base
    
    for ded_name, ded_type, amount, is_percentage, is_pre_tax in deductions:
        if is_pre_tax:
            if is_percentage:
                ded_amount = gross * (amount / 100)
            else:
                ded_amount = amount
            pre_tax_deductions[ded_name] = ded_amount
            pre_tax_total += ded_amount
            # 401k reduces Social Security taxable base, health insurance does not
            if ded_type == '401k':
                ss_taxable_reduction += ded_amount
    
    # Calculate taxable income for federal/state tax (gross minus all pre-tax deductions)
    taxable_income = max(0, gross - pre_tax_total)
    
    # Calculate SS taxable base (gross minus 401k only)
    ss_taxable_base = max(0, gross - ss_taxable_reduction)
    
    # Calculate taxes
    pay_frequency = employee_data.get('pay_frequency', 'biweekly')
    # Federal tax is calculated per period (standard deduction is divided across periods)
    fed_tax = calculate_federal_tax(taxable_income, ytd_gross=0, pay_frequency=pay_frequency)
    state_tax = calculate_state_tax(taxable_income, state)
    # Social Security is calculated on gross minus 401k contributions only
    ss, medicare = calculate_fica(ss_taxable_base, ytd_gross=ytd_gross)
    
    # Calculate post-tax deductions
    post_tax_deductions = {}
    post_tax_total = 0
    
    for ded_name, ded_type, amount, is_percentage, is_pre_tax in deductions:
        if not is_pre_tax:
            if is_percentage:
                ded_amount = gross * (amount / 100)
            else:
                ded_amount = amount
            post_tax_deductions[ded_name] = ded_amount
            post_tax_total += ded_amount
    
    # Calculate net pay
    net = gross - pre_tax_total - fed_tax - state_tax - ss - medicare - post_tax_total
    
    # Combine all deductions for return
    all_deductions = {**pre_tax_deductions, **post_tax_deductions}
    
    return {
        'gross': round(gross, 2),
        'federal_tax': round(fed_tax, 2),
        'state_tax': round(state_tax, 2),
        'social_security': round(ss, 2),
        'medicare': round(medicare, 2),
        'total_fica': round(ss + medicare, 2),
        'pre_tax_deductions': pre_tax_deductions,
        'post_tax_deductions': post_tax_deductions,
        'total_pre_tax_deductions': round(pre_tax_total, 2),
        'total_post_tax_deductions': round(post_tax_total, 2),
        'other_deductions': round(pre_tax_total + post_tax_total, 2),
        'net': round(net, 2)
    }

# Transaction generation data - location-based merchants
LOCATION_MERCHANTS = {
    'CA': {
        'groceries': ['Whole Foods', 'Trader Joe\'s', 'Safeway', 'Ralphs', 'Vons', 'Albertsons'],
        'gas': ['Chevron', 'Shell', '76', 'ARCO', 'Mobil'],
        'restaurants': ['In-N-Out Burger', 'Chipotle', 'Starbucks', 'McDonald\'s', 'Taco Bell', 'Panda Express'],
        'retail': ['Target', 'Walmart', 'Costco', 'Best Buy', 'Home Depot'],
        'utilities': ['PG&E', 'SoCal Edison', 'SDG&E'],
        'entertainment': ['AMC Theatres', 'Regal Cinemas', 'Netflix', 'Spotify']
    },
    'NY': {
        'groceries': ['Whole Foods', 'Trader Joe\'s', 'Stop & Shop', 'Key Food', 'Fairway'],
        'gas': ['Shell', 'Mobil', 'BP', 'Exxon', 'Sunoco'],
        'restaurants': ['Shake Shack', 'Chipotle', 'Starbucks', 'McDonald\'s', 'Dunkin\'', 'Subway'],
        'retail': ['Target', 'Walmart', 'Best Buy', 'Home Depot', 'Macy\'s'],
        'utilities': ['Con Edison', 'National Grid', 'PSEG'],
        'entertainment': ['AMC Theatres', 'Regal Cinemas', 'Netflix', 'Spotify']
    },
    'TX': {
        'groceries': ['H-E-B', 'Kroger', 'Walmart', 'Whole Foods', 'Randalls'],
        'gas': ['Exxon', 'Shell', 'Chevron', 'Valero', '7-Eleven'],
        'restaurants': ['Whataburger', 'Chipotle', 'Starbucks', 'McDonald\'s', 'Taco Bell', 'Chick-fil-A'],
        'retail': ['Target', 'Walmart', 'Best Buy', 'Home Depot', 'Lowe\'s'],
        'utilities': ['TXU Energy', 'Reliant', 'Oncor'],
        'entertainment': ['AMC Theatres', 'Cinemark', 'Netflix', 'Spotify']
    },
    'FL': {
        'groceries': ['Publix', 'Winn-Dixie', 'Walmart', 'Whole Foods', 'Aldi'],
        'gas': ['Shell', 'Chevron', 'Exxon', 'BP', '7-Eleven'],
        'restaurants': ['Pollo Tropical', 'Chipotle', 'Starbucks', 'McDonald\'s', 'Subway', 'Papa John\'s'],
        'retail': ['Target', 'Walmart', 'Best Buy', 'Home Depot', 'Lowe\'s'],
        'utilities': ['FPL', 'Duke Energy', 'TECO'],
        'entertainment': ['AMC Theatres', 'Regal Cinemas', 'Netflix', 'Spotify']
    }
}

# Default merchants if state not found
DEFAULT_MERCHANTS = {
    'groceries': ['Walmart', 'Kroger', 'Target', 'Whole Foods', 'Safeway'],
    'gas': ['Shell', 'Chevron', 'Exxon', 'BP', 'Mobil'],
    'restaurants': ['McDonald\'s', 'Starbucks', 'Chipotle', 'Subway', 'Taco Bell'],
    'retail': ['Target', 'Walmart', 'Best Buy', 'Home Depot', 'Lowe\'s'],
    'utilities': ['Electric Company', 'Gas Company', 'Water Company'],
    'entertainment': ['AMC Theatres', 'Regal Cinemas', 'Netflix', 'Spotify']
}

TRANSACTION_CATEGORIES = {
    'groceries': {'min': 25, 'max': 150, 'frequency': 0.20},
    'gas': {'min': 30, 'max': 80, 'frequency': 0.12},
    'restaurants': {'min': 8, 'max': 45, 'frequency': 0.25},
    'utilities': {'min': 50, 'max': 200, 'frequency': 0.05},
    'retail': {'min': 15, 'max': 200, 'frequency': 0.15},
    'entertainment': {'min': 10, 'max': 50, 'frequency': 0.10},
    'subscriptions': {'min': 5, 'max': 20, 'frequency': 0.05},
    'pharmacy': {'min': 10, 'max': 80, 'frequency': 0.05},
    'atm': {'min': 20, 'max': 100, 'frequency': 0.03}
}

def get_merchants_for_location(state):
    """Get merchants based on state location"""
    return LOCATION_MERCHANTS.get(state, DEFAULT_MERCHANTS)

def generate_transactions(paystub_id, employee_id, user_id, net_pay, period_start, period_end, city, state):
    """Generate 45-70 realistic transactions for a pay period"""
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Get merchants for this location
    merchants = get_merchants_for_location(state)
    
    # Calculate number of transactions (45-70)
    num_transactions = random.randint(45, 70)
    
    # Generate date range for transactions
    start_date = datetime.strptime(period_start, '%Y-%m-%d').date() if isinstance(period_start, str) else period_start
    end_date = datetime.strptime(period_end, '%Y-%m-%d').date() if isinstance(period_end, str) else period_end
    date_range = (end_date - start_date).days + 1
    
    transactions = []
    
    # Add paystub deposit as first transaction
    deposit_date = end_date  # Payday is typically on period_end
    transactions.append({
        'paystub_id': paystub_id,
        'employee_id': employee_id,
        'user_id': user_id,
        'transaction_date': deposit_date,
        'description': 'Direct Deposit - Payroll',
        'merchant': 'EMPLOYER PAYROLL',
        'category': 'income',
        'amount': net_pay,
        'transaction_type': 'deposit',
        'location_city': city,
        'location_state': state
    })
    
    # Generate spending transactions
    remaining_transactions = num_transactions - 1
    
    # Distribute transactions across the date range
    for i in range(remaining_transactions):
        # Random date within period
        days_offset = random.randint(0, date_range - 1)
        transaction_date = start_date + timedelta(days=days_offset)
        
        # Select category based on frequency
        rand = random.random()
        cumulative = 0
        selected_category = 'groceries'  # default
        
        for cat, info in TRANSACTION_CATEGORIES.items():
            cumulative += info['frequency']
            if rand <= cumulative:
                selected_category = cat
                break
        
        # Get amount for category
        cat_info = TRANSACTION_CATEGORIES[selected_category]
        amount = round(random.uniform(cat_info['min'], cat_info['max']), 2)
        
        # Get merchant for category
        if selected_category in merchants:
            merchant = random.choice(merchants[selected_category])
        else:
            merchant = random.choice(DEFAULT_MERCHANTS.get(selected_category, ['Merchant']))
        
        # Special handling for certain categories
        if selected_category == 'subscriptions':
            merchant = random.choice(['Netflix', 'Spotify', 'Amazon Prime', 'Disney+', 'Hulu', 'Apple Music'])
        elif selected_category == 'pharmacy':
            merchant = random.choice(['CVS', 'Walgreens', 'Rite Aid'])
        elif selected_category == 'atm':
            merchant = 'ATM WITHDRAWAL'
            amount = round(random.choice([20, 40, 60, 80, 100]), 2)
        
        # Create description
        if selected_category == 'atm':
            description = f'ATM Withdrawal - {city}, {state}'
        else:
            description = f'{merchant} - {city}, {state}'
        
        transactions.append({
            'paystub_id': paystub_id,
            'employee_id': employee_id,
            'user_id': user_id,
            'transaction_date': transaction_date,
            'description': description,
            'merchant': merchant,
            'category': selected_category,
            'amount': amount,
            'transaction_type': 'debit',
            'location_city': city,
            'location_state': state
        })
    
    # Sort transactions by date
    transactions.sort(key=lambda x: x['transaction_date'])
    
    # Insert transactions into database
    for trans in transactions:
        c.execute("""
            INSERT INTO transactions (
                paystub_id, employee_id, user_id, transaction_date,
                description, merchant, category, amount, transaction_type,
                location_city, location_state
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            trans['paystub_id'], trans['employee_id'], trans['user_id'],
            trans['transaction_date'], trans['description'], trans['merchant'],
            trans['category'], trans['amount'], trans['transaction_type'],
            trans['location_city'], trans['location_state']
        ))
    
    conn.commit()
    conn.close()
    
    return len(transactions)

def get_paystub_ytd(employee_id, current_check_number):
    """Get YTD totals up to current check number"""
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("""
        SELECT 
            SUM(gross_pay) as ytd_gross,
            SUM(federal_tax) as ytd_federal,
            SUM(state_tax) as ytd_state,
            SUM(social_security + medicare) as ytd_fica,
            SUM(net_pay) as ytd_net
        FROM paystubs 
        WHERE employee_id=? AND check_number < ?
    """, (employee_id, current_check_number))
    
    result = c.fetchone()
    conn.close()
    
    if result and result[0]:
        return {
            'ytd_gross': result[0],
            'ytd_federal': result[1],
            'ytd_state': result[2],
            'ytd_fica': result[3],
            'ytd_net': result[4]
        }
    return {
        'ytd_gross': 0, 'ytd_federal': 0, 'ytd_state': 0,
        'ytd_fica': 0, 'ytd_net': 0
    }

def generate_paystub_pdf(employee, paystub, company):
    """Generate ADP-style paystub PDF"""
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Header - Company Info
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, str(company.get('name', '')))
    
    p.setFont("Helvetica", 9)
    p.drawString(50, height - 65, str(company.get('address', '')))
    p.drawString(50, height - 75, f"EIN: {str(company.get('ein', ''))}")
    
    # Check info
    p.setFont("Helvetica-Bold", 10)
    p.drawString(400, height - 50, f"CHECK #{paystub['check_number']}")
    p.drawString(400, height - 65, f"PAY DATE: {paystub['period_end']}")
    
    # Employee Info Panel
    p.setFont("Helvetica", 9)
    p.drawString(50, height - 110, "EMPLOYEE")
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, height - 125, str(employee.get('name', '')))
    p.setFont("Helvetica", 9)
    # Safely handle SSN - ensure it's a string and has at least 4 characters
    ssn = str(employee.get('ssn', ''))
    ssn_display = f"***-**-{ssn[-4:]}" if len(ssn) >= 4 else "***-**-****"
    p.drawString(50, height - 140, f"SSN: {ssn_display}")
    p.drawString(50, height - 155, str(employee.get('street', '')))
    p.drawString(50, height - 170, f"{employee.get('city', '')}, {employee.get('state', '')} {employee.get('zip', '')}")
    
    # Earnings Section
    y = height - 220
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "EARNINGS")
    p.setFont("Helvetica", 9)
    
    # Column headers
    p.drawString(50, y - 15, "Description")
    p.drawString(200, y - 15, "Hours")
    p.drawString(280, y - 15, "Rate")
    p.drawString(340, y - 15, "This Period")
    p.drawString(450, y - 15, "Year to Date")
    
    # Earnings rows
    p.drawString(50, y - 30, "Regular Pay")
    if employee['is_hourly']:
        p.drawString(200, y - 30, f"{paystub.get('hours_worked', 80):.1f}")
    p.drawString(280, y - 30, f"${employee['pay_rate']:.2f}")
    p.drawString(340, y - 30, f"${paystub['gross_pay']:.2f}")
    p.drawString(450, y - 30, f"${paystub['ytd_gross']:.2f}")
    
    # Deductions Section
    y -= 80
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "DEDUCTIONS")
    p.setFont("Helvetica", 9)
    
    p.drawString(50, y - 15, "Description")
    p.drawString(340, y - 15, "This Period")
    p.drawString(450, y - 15, "Year to Date")
    
    deductions_y = y - 30
    p.drawString(50, deductions_y, "Federal Income Tax")
    p.drawString(340, deductions_y, f"${paystub['federal_tax']:.2f}")
    p.drawString(450, deductions_y, f"${paystub['ytd_federal']:.2f}")
    
    deductions_y -= 15
    p.drawString(50, deductions_y, "State Income Tax")
    p.drawString(340, deductions_y, f"${paystub['state_tax']:.2f}")
    p.drawString(450, deductions_y, f"${paystub['ytd_state']:.2f}")
    
    deductions_y -= 15
    p.drawString(50, deductions_y, "Social Security (6.2%)")
    p.drawString(340, deductions_y, f"${paystub['social_security']:.2f}")
    p.drawString(450, deductions_y, f"${paystub['ytd_fica']:.2f}")
    
    deductions_y -= 15
    p.drawString(50, deductions_y, "Medicare (1.45%)")
    p.drawString(340, deductions_y, f"${paystub['medicare']:.2f}")
    
    # Custom deductions (pre-tax and post-tax)
    pre_tax_deductions = paystub.get('pre_tax_deductions', {})
    post_tax_deductions = paystub.get('post_tax_deductions', {})
    
    # Add pre-tax deductions
    for ded_name, ded_amount in pre_tax_deductions.items():
        deductions_y -= 15
        p.drawString(50, deductions_y, ded_name)
        p.drawString(340, deductions_y, f"${ded_amount:.2f}")
    
    # Add post-tax deductions
    for ded_name, ded_amount in post_tax_deductions.items():
        deductions_y -= 15
        p.drawString(50, deductions_y, ded_name)
        p.drawString(340, deductions_y, f"${ded_amount:.2f}")
    
    # Net Pay - Highlighted
    deductions_y -= 35
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, deductions_y, "NET PAY")
    p.drawString(340, deductions_y, f"${paystub['net_pay']:.2f}")
    
    # Summary line at bottom
    p.setFont("Helvetica", 8)
    p.drawString(50, 40, "This is an electronically generated document. No signature required.")
    
    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer.getvalue()

# Routes

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('profile'))
    return render_template('landing.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        confirm = request.form['confirm']
        
        if password != confirm:
            flash('Passwords do not match', 'error')
            return redirect(url_for('register'))
        
        conn = sqlite3.connect('paystub.db')
        c = conn.cursor()
        try:
            hashed = generate_password_hash(password)
            c.execute("INSERT INTO users (email, password) VALUES (?, ?)", (email, hashed))
            conn.commit()
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Email already registered', 'error')
        finally:
            conn.close()
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = sqlite3.connect('paystub.db')
        c = conn.cursor()
        c.execute("SELECT id, email, password FROM users WHERE email=?", (email,))
        user = c.fetchone()
        conn.close()
        
        if user and check_password_hash(user[2], password):
            login_user(User(user[0], user[1]))
            return redirect(url_for('profile'))
        
        flash('Invalid email or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/profile')
@login_required
def profile():
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Get companies
    c.execute("SELECT id, name, ein FROM companies WHERE user_id=? ORDER BY name", 
              (current_user.id,))
    companies = c.fetchall()
    
    # Get employees
    c.execute("""
        SELECT e.id, e.name, e.city, e.state, c.name as company_name
        FROM employees e
        JOIN companies c ON e.company_id = c.id
        WHERE e.user_id = ?
        ORDER BY c.name, e.name
    """, (current_user.id,))
    employees = c.fetchall()
    
    # Get recent paystubs
    c.execute("""
        SELECT p.id, p.check_number, p.period_start, p.period_end, p.gross_pay, 
               p.net_pay, e.name, p.edited
        FROM paystubs p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.user_id = ?
        ORDER BY p.check_number DESC
        LIMIT 20
    """, (current_user.id,))
    paystubs = c.fetchall()
    
    conn.close()
    
    return render_template('profile.html', companies=companies, employees=employees, paystubs=paystubs)

@app.route('/company/create', methods=['GET', 'POST'])
@login_required
def create_company():
    if request.method == 'POST':
        name = request.form['name']
        ein = request.form['ein']
        address = request.form['address']
        phone = request.form['phone']
        
        conn = sqlite3.connect('paystub.db')
        c = conn.cursor()
        c.execute("""
            INSERT INTO companies (user_id, name, ein, address, phone)
            VALUES (?, ?, ?, ?, ?)
        """, (current_user.id, name, ein, address, phone))
        conn.commit()
        conn.close()
        
        flash('Company created successfully!', 'success')
        return redirect(url_for('profile'))
    
    return render_template('create_company.html')

@app.route('/employee/create', methods=['GET', 'POST'])
@login_required
def create_employee():
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("SELECT id, name FROM companies WHERE user_id=?", (current_user.id,))
    companies = c.fetchall()
    
    if request.method == 'POST':
        company_id = request.form['company_id']
        name = request.form['name']
        ssn = request.form['ssn']
        street = request.form['street']
        city = request.form['city']
        state = request.form['state']
        zip_code = request.form['zip']
        pay_rate = float(request.form['pay_rate'])
        is_hourly = 1 if request.form['pay_type'] == 'hourly' else 0
        pay_frequency = request.form['pay_frequency']
        
        c.execute("""
            INSERT INTO employees (user_id, company_id, name, ssn, street, 
                                   city, state, zip, pay_rate, is_hourly, pay_frequency)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (current_user.id, company_id, name, ssn, street, city, 
              state, zip_code, pay_rate, is_hourly, pay_frequency))
        employee_id = c.lastrowid
        
        # Process deductions
        deduction_names = request.form.getlist('deduction_name[]')
        deduction_types = request.form.getlist('deduction_type[]')
        deduction_amounts = request.form.getlist('deduction_amount[]')
        deduction_is_percentage = request.form.getlist('deduction_is_percentage[]')
        deduction_is_pre_tax = request.form.getlist('deduction_is_pre_tax[]')
        
        for i, ded_name in enumerate(deduction_names):
            if ded_name and ded_name.strip():
                try:
                    amount = float(deduction_amounts[i]) if i < len(deduction_amounts) else 0
                    is_percentage = 1 if (i < len(deduction_is_percentage) and deduction_is_percentage[i] == '1') else 0
                    is_pre_tax = 1 if (i < len(deduction_is_pre_tax) and deduction_is_pre_tax[i] == '1') else 0
                    ded_type = deduction_types[i] if i < len(deduction_types) else 'other'
                    
                    c.execute("""
                        INSERT INTO employee_deductions (employee_id, user_id, deduction_name, deduction_type, amount, is_percentage, is_pre_tax)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (employee_id, current_user.id, ded_name.strip(), ded_type, amount, is_percentage, is_pre_tax))
                except (ValueError, IndexError):
                    continue
        
        conn.commit()
        conn.close()
        
        flash('Employee created successfully!', 'success')
        return redirect(url_for('profile'))
    
    conn.close()
    return render_template('create_employee.html', companies=companies)

@app.route('/employee/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Verify employee belongs to user
    c.execute("SELECT * FROM employees WHERE id=? AND user_id=?", 
             (employee_id, current_user.id))
    employee = c.fetchone()
    
    if not employee:
        flash('Employee not found', 'error')
        return redirect(url_for('profile'))
    
    # Get companies for dropdown
    c.execute("SELECT id, name FROM companies WHERE user_id=?", (current_user.id,))
    companies = c.fetchall()
    
    if request.method == 'POST':
        company_id = request.form['company_id']
        name = request.form['name']
        ssn = request.form['ssn']
        street = request.form['street']
        city = request.form['city']
        state = request.form['state']
        zip_code = request.form['zip']
        pay_rate = float(request.form['pay_rate'])
        is_hourly = 1 if request.form['pay_type'] == 'hourly' else 0
        pay_frequency = request.form['pay_frequency']
        
        c.execute("""
            UPDATE employees 
            SET company_id=?, name=?, ssn=?, street=?, city=?, state=?, zip=?,
                pay_rate=?, is_hourly=?, pay_frequency=?
            WHERE id=? AND user_id=?
        """, (company_id, name, ssn, street, city, state, zip_code,
              pay_rate, is_hourly, pay_frequency, employee_id, current_user.id))
        
        # Delete existing deductions and recreate them
        c.execute("DELETE FROM employee_deductions WHERE employee_id=? AND user_id=?", (employee_id, current_user.id))
        
        # Process deductions
        deduction_names = request.form.getlist('deduction_name[]')
        deduction_types = request.form.getlist('deduction_type[]')
        deduction_amounts = request.form.getlist('deduction_amount[]')
        deduction_is_percentage = request.form.getlist('deduction_is_percentage[]')
        deduction_is_pre_tax = request.form.getlist('deduction_is_pre_tax[]')
        
        for i, ded_name in enumerate(deduction_names):
            if ded_name and ded_name.strip():
                try:
                    amount = float(deduction_amounts[i]) if i < len(deduction_amounts) else 0
                    is_percentage = 1 if (i < len(deduction_is_percentage) and deduction_is_percentage[i] == '1') else 0
                    is_pre_tax = 1 if (i < len(deduction_is_pre_tax) and deduction_is_pre_tax[i] == '1') else 0
                    ded_type = deduction_types[i] if i < len(deduction_types) else 'other'
                    
                    c.execute("""
                        INSERT INTO employee_deductions (employee_id, user_id, deduction_name, deduction_type, amount, is_percentage, is_pre_tax)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (employee_id, current_user.id, ded_name.strip(), ded_type, amount, is_percentage, is_pre_tax))
                except (ValueError, IndexError):
                    continue
        
        conn.commit()
        conn.close()
        
        flash('Employee updated successfully!', 'success')
        return redirect(url_for('profile'))
    
    # Get existing deductions
    c.execute("""
        SELECT id, deduction_name, deduction_type, amount, is_percentage, is_pre_tax
        FROM employee_deductions
        WHERE employee_id=? AND user_id=?
        ORDER BY is_pre_tax DESC, deduction_name
    """, (employee_id, current_user.id))
    deductions = c.fetchall()
    
    # Prepare employee data for form
    # Employee structure: id, user_id, company_id, name, ssn, street, city, state, zip, pay_rate, is_hourly, pay_frequency, created_at
    employee_data = {
        'id': employee[0],
        'company_id': employee[2],
        'name': employee[3],
        'ssn': employee[4] or '',
        'street': employee[5] or '',
        'city': employee[6] or '',
        'state': employee[7] or '',
        'zip': employee[8] or '',
        'pay_rate': employee[9],
        'is_hourly': bool(employee[10]),
        'pay_frequency': employee[11] or 'biweekly'
    }
    
    conn.close()
    return render_template('edit_employee.html', employee=employee_data, companies=companies, deductions=deductions)

@app.route('/employee/delete/<int:employee_id>')
@login_required
def delete_employee(employee_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Verify employee belongs to user
    c.execute("SELECT id FROM employees WHERE id=? AND user_id=?", 
             (employee_id, current_user.id))
    employee = c.fetchone()
    
    if not employee:
        flash('Employee not found', 'error')
        return redirect(url_for('profile'))
    
    # Delete employee (cascade will delete paystubs and transactions)
    c.execute("DELETE FROM employees WHERE id=? AND user_id=?", 
             (employee_id, current_user.id))
    
    conn.commit()
    conn.close()
    
    flash('Employee deleted successfully!', 'success')
    return redirect(url_for('profile'))

@app.route('/generate', methods=['GET', 'POST'])
@login_required
def generate():
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("""
        SELECT e.id, e.name, e.company_id, c.name as company_name
        FROM employees e
        JOIN companies c ON e.company_id = c.id
        WHERE e.user_id = ?
        ORDER BY c.name, e.name
    """, (current_user.id,))
    employees = c.fetchall()
    
    if request.method == 'POST':
        employee_id = int(request.form['employee_id'])
        start_check_num = int(request.form['start_check_num'])
        num_stubs = int(request.form['num_stubs'])
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        direction = request.form['direction']  # 'future' or 'past'
        
        # Get employee data
        c.execute("SELECT * FROM employees WHERE id=? AND user_id=?", 
                 (employee_id, current_user.id))
        emp = c.fetchone()
        emp_dict = {
            'id': emp[0], 
            'name': emp[3],  # Fixed: was emp[2]
            'ssn': str(emp[4]) if emp[4] else '',  # Fixed: was emp[5], ensure string
            'street': str(emp[5]) if emp[5] else '',  # Added missing field
            'city': str(emp[6]) if emp[6] else '',  # Added missing field
            'state': str(emp[7]) if emp[7] else '',  # Ensure string
            'zip': str(emp[8]) if emp[8] else '',  # Added missing field
            'pay_rate': float(emp[9]), 
            'is_hourly': bool(emp[10]),
            'pay_frequency': str(emp[11]) if emp[11] else 'biweekly'
        }
        
        # Get company data
        c.execute("SELECT * FROM companies WHERE id=?", (emp[2],))
        company = c.fetchone()
        company_dict = {
            'id': company[0], 
            'name': str(company[2]) if company[2] else '', 
            'ein': str(company[3]) if company[3] else '',
            'address': str(company[4]) if company[4] else ''
        }
        
        # Calculate pay period days
        freq_map = {'weekly': 7, 'biweekly': 14, 'semimonthly': 15, 'monthly': 30}
        period_days = freq_map.get(emp_dict['pay_frequency'], 14)
        
        # Generate stubs
        for i in range(num_stubs):
            check_num = start_check_num + (i if direction == 'future' else -i)
            
            if direction == 'future':
                period_start = start_date + timedelta(days=i * period_days)
            else:
                period_start = start_date - timedelta(days=i * period_days)
            
            period_end = period_start + timedelta(days=period_days - 1)
            
            # Calculate hours for hourly
            hours = 80 if emp_dict['is_hourly'] else 0
            
            # Get YTD before this period (for accurate tax calculations)
            ytd = get_paystub_ytd(employee_id, check_num)
            
            # Calculate pay (pass YTD gross for accurate Social Security calculation)
            pay_calc = calculate_paystub(emp_dict, hours, ytd_gross=ytd['ytd_gross'])
            
            # Add current period to YTD
            ytd_gross = ytd['ytd_gross'] + pay_calc['gross']
            ytd_federal = ytd['ytd_federal'] + pay_calc['federal_tax']
            ytd_state = ytd['ytd_state'] + pay_calc['state_tax']
            ytd_fica = ytd['ytd_fica'] + pay_calc['total_fica']
            ytd_net = ytd['ytd_net'] + pay_calc['net']
            
            # Generate PDF
            paystub_data = {
                'check_number': check_num,
                'period_start': str(period_start),
                'period_end': str(period_end),
                'hours_worked': hours,
                'gross_pay': pay_calc['gross'],
                'federal_tax': pay_calc['federal_tax'],
                'state_tax': pay_calc['state_tax'],
                'social_security': pay_calc['social_security'],
                'medicare': pay_calc['medicare'],
                'net_pay': pay_calc['net'],
                'ytd_gross': ytd_gross,
                'ytd_federal': ytd_federal,
                'ytd_state': ytd_state,
                'ytd_fica': ytd_fica,
                'ytd_net': ytd_net,
                'pre_tax_deductions': pay_calc['pre_tax_deductions'],
                'post_tax_deductions': pay_calc['post_tax_deductions']
            }
            
            pdf_blob = generate_paystub_pdf(emp_dict, paystub_data, company_dict)
            
            # Save to database
            c.execute("""
                INSERT INTO paystubs (
                    employee_id, user_id, check_number, period_start, period_end,
                    hours_worked, gross_pay, federal_tax, state_tax,
                    social_security, medicare, other_deductions, net_pay,
                    ytd_gross, ytd_federal, ytd_state, ytd_fica, ytd_net,
                    pdf_blob
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                employee_id, current_user.id, check_num, period_start, period_end,
                hours, pay_calc['gross'], pay_calc['federal_tax'], pay_calc['state_tax'],
                pay_calc['social_security'], pay_calc['medicare'], pay_calc['other_deductions'], pay_calc['net'],
                ytd_gross, ytd_federal, ytd_state, ytd_fica, ytd_net,
                pdf_blob
            ))
            
            # Get the paystub ID
            paystub_id = c.lastrowid
            conn.commit()  # Commit paystub first
            
            # Generate transactions for this paystub
            generate_transactions(
                paystub_id=paystub_id,
                employee_id=employee_id,
                user_id=current_user.id,
                net_pay=pay_calc['net'],
                period_start=period_start,
                period_end=period_end,
                city=emp_dict.get('city', ''),
                state=emp_dict.get('state', '')
            )
        
        conn.close()
        
        flash(f'Generated {num_stubs} paystubs successfully!', 'success')
        return redirect(url_for('profile'))
    
    conn.close()
    return render_template('generate.html', employees=employees)

@app.route('/edit/<int:stub_id>', methods=['GET', 'POST'])
@login_required
def edit_stub(stub_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    c.execute("""
        SELECT * FROM paystubs WHERE id=? AND user_id=?
    """, (stub_id, current_user.id))
    paystub = c.fetchone()
    
    if not paystub:
        flash('Paystub not found', 'error')
        return redirect(url_for('profile'))
    
    if request.method == 'POST':
        gross_pay = float(request.form.get('gross_pay', paystub[7]))
        federal_tax = float(request.form.get('federal_tax', paystub[8]))
        state_tax = float(request.form.get('state_tax', paystub[9]))
        propagate = request.form.get('propagate', False)
        
        net_pay = gross_pay - federal_tax - state_tax - paystub[10] - paystub[11]
        
        # Update this stub
        c.execute("""
            UPDATE paystubs 
            SET gross_pay=?, federal_tax=?, state_tax=?, net_pay=?, edited=1
            WHERE id=?
        """, (gross_pay, federal_tax, state_tax, net_pay, stub_id))
        
        # Record edit
        c.execute("""
            INSERT INTO stub_edits (paystub_id, field_name, old_value, new_value, propagate_to_later)
            VALUES (?, ?, ?, ?, ?)
        """, (stub_id, 'gross_pay', str(paystub[7]), str(gross_pay), 1 if propagate else 0))
        
        # Propagate to later stubs if requested
        if propagate:
            c.execute("""
                SELECT id, gross_pay FROM paystubs 
                WHERE employee_id=? AND check_number > ? AND user_id=?
                ORDER BY check_number
            """, (paystub[1], paystub[3], current_user.id))
            later_stubs = c.fetchall()
            
            gross_diff = gross_pay - paystub[7]
            
            for later_id, later_gross in later_stubs:
                new_later_gross = later_gross + gross_diff
                c.execute("""
                    UPDATE paystubs 
                    SET gross_pay=?, ytd_gross=ytd_gross+?, edited=1
                    WHERE id=?
                """, (new_later_gross, gross_diff, later_id))
        
        conn.commit()
        conn.close()
        
        flash('Paystub updated successfully!', 'success')
        return redirect(url_for('profile'))
    
    conn.close()
    return render_template('edit.html', paystub=paystub)

@app.route('/view/<int:stub_id>')
@login_required
def view_stub(stub_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    c.execute("""
        SELECT p.*, e.name, e.ssn, e.street, e.city, e.state, e.zip, c.name as company_name
        FROM paystubs p
        JOIN employees e ON p.employee_id = e.id
        JOIN companies c ON e.company_id = c.id
        WHERE p.id=? AND p.user_id=?
    """, (stub_id, current_user.id))
    
    stub = c.fetchone()
    conn.close()
    
    if not stub:
        flash('Paystub not found', 'error')
        return redirect(url_for('profile'))
    
    return render_template('view.html', stub=stub)

@app.route('/download/<int:stub_id>')
@login_required
def download_stub(stub_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("SELECT pdf_blob FROM paystubs WHERE id=? AND user_id=?", (stub_id, current_user.id))
    result = c.fetchone()
    conn.close()
    
    if not result or not result[0]:
        flash('PDF not found', 'error')
        return redirect(url_for('profile'))
    
    return send_file(
        io.BytesIO(result[0]),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'paystub_{stub_id}.pdf'
    )

@app.route('/delete/<int:stub_id>')
@login_required
def delete_stub(stub_id):
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    c.execute("DELETE FROM paystubs WHERE id=? AND user_id=?", (stub_id, current_user.id))
    conn.commit()
    conn.close()
    
    flash('Paystub deleted successfully!', 'success')
    return redirect(url_for('profile'))

@app.route('/delete/mass', methods=['POST'])
@login_required
def mass_delete_stubs():
    """Mass delete multiple paystubs"""
    stub_ids = request.form.getlist('stub_ids')
    
    if not stub_ids:
        flash('No paystubs selected', 'error')
        return redirect(url_for('profile'))
    
    # Convert to integers and validate
    try:
        stub_ids = [int(id) for id in stub_ids]
    except ValueError:
        flash('Invalid paystub IDs', 'error')
        return redirect(url_for('profile'))
    
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Verify all paystubs belong to user and delete
    placeholders = ','.join('?' * len(stub_ids))
    c.execute(f"""
        DELETE FROM paystubs 
        WHERE id IN ({placeholders}) AND user_id=?
    """, stub_ids + [current_user.id])
    
    deleted_count = c.rowcount
    conn.commit()
    conn.close()
    
    flash(f'Successfully deleted {deleted_count} paystub(s)!', 'success')
    return redirect(url_for('profile'))

@app.route('/transactions/<int:stub_id>')
@login_required
def view_transactions(stub_id):
    """View transactions for a specific paystub"""
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Verify paystub belongs to user
    c.execute("""
        SELECT p.id, p.check_number, p.period_start, p.period_end, p.net_pay,
               e.name, e.city, e.state
        FROM paystubs p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.id=? AND p.user_id=?
    """, (stub_id, current_user.id))
    
    paystub_info = c.fetchone()
    
    if not paystub_info:
        flash('Paystub not found', 'error')
        return redirect(url_for('profile'))
    
    # Get transactions
    c.execute("""
        SELECT id, transaction_date, description, merchant, category,
               amount, transaction_type, location_city, location_state
        FROM transactions
        WHERE paystub_id=? AND user_id=?
        ORDER BY transaction_date, id
    """, (stub_id, current_user.id))
    
    transactions = c.fetchall()
    
    # Calculate totals
    total_deposits = sum(t[5] for t in transactions if t[6] == 'deposit')
    total_debits = sum(t[5] for t in transactions if t[6] == 'debit')
    balance = total_deposits - total_debits
    
    # Calculate category totals
    category_totals = {}
    for trans in transactions:
        if trans[6] == 'debit':  # Only count debits
            cat = trans[4]  # category
            if cat not in category_totals:
                category_totals[cat] = 0
            category_totals[cat] += trans[5]
    
    conn.close()
    
    return render_template('transactions.html', 
                         paystub_info=paystub_info,
                         transactions=transactions,
                         total_deposits=total_deposits,
                         total_debits=total_debits,
                         balance=balance,
                         category_totals=category_totals)

@app.route('/transactions/generate/<int:stub_id>')
@login_required
def generate_transactions_route(stub_id):
    """Manually generate transactions for a paystub"""
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Verify paystub belongs to user and get info
    c.execute("""
        SELECT p.id, p.employee_id, p.net_pay, p.period_start, p.period_end,
               e.city, e.state, e.user_id
        FROM paystubs p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.id=? AND p.user_id=?
    """, (stub_id, current_user.id))
    
    paystub = c.fetchone()
    
    if not paystub:
        flash('Paystub not found', 'error')
        return redirect(url_for('profile'))
    
    # Check if transactions already exist
    c.execute("SELECT COUNT(*) FROM transactions WHERE paystub_id=?", (stub_id,))
    existing_count = c.fetchone()[0]
    
    if existing_count > 0:
        flash('Transactions already exist for this paystub. Delete them first to regenerate.', 'error')
        return redirect(url_for('view_transactions', stub_id=stub_id))
    
    # Generate transactions
    num_transactions = generate_transactions(
        paystub_id=paystub[0],
        employee_id=paystub[1],
        user_id=paystub[7],
        net_pay=paystub[2],
        period_start=paystub[3],
        period_end=paystub[4],
        city=paystub[5] or '',
        state=paystub[6] or ''
    )
    
    conn.close()
    
    flash(f'Generated {num_transactions} transactions successfully!', 'success')
    return redirect(url_for('view_transactions', stub_id=stub_id))

@app.route('/transactions/export/<int:stub_id>')
@login_required
def export_transactions_excel(stub_id):
    """Export transactions to Excel file"""
    conn = sqlite3.connect('paystub.db')
    c = conn.cursor()
    
    # Verify paystub belongs to user
    c.execute("""
        SELECT p.id, p.check_number, p.period_start, p.period_end, p.net_pay,
               e.name, e.city, e.state
        FROM paystubs p
        JOIN employees e ON p.employee_id = e.id
        WHERE p.id=? AND p.user_id=?
    """, (stub_id, current_user.id))
    
    paystub_info = c.fetchone()
    
    if not paystub_info:
        flash('Paystub not found', 'error')
        return redirect(url_for('profile'))
    
    # Get transactions
    c.execute("""
        SELECT id, transaction_date, description, merchant, category,
               amount, transaction_type, location_city, location_state
        FROM transactions
        WHERE paystub_id=? AND user_id=?
        ORDER BY transaction_date, id
    """, (stub_id, current_user.id))
    
    transactions = c.fetchall()
    
    if not transactions:
        flash('No transactions to export', 'error')
        return redirect(url_for('view_transactions', stub_id=stub_id))
    
    # Calculate totals
    total_deposits = sum(t[5] for t in transactions if t[6] == 'deposit')
    total_debits = sum(t[5] for t in transactions if t[6] == 'debit')
    balance = total_deposits - total_debits
    
    # Calculate category totals
    category_totals = {}
    for trans in transactions:
        if trans[6] == 'debit':
            cat = trans[4]
            if cat not in category_totals:
                category_totals[cat] = 0
            category_totals[cat] += trans[5]
    
    conn.close()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Define styles
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    currency_font = Font(size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title and summary information
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Transaction History - Check #{paystub_info[1]}"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Employee: {paystub_info[5]} | Period: {paystub_info[2]} to {paystub_info[3]}"
    ws[f'A{row}'].font = Font(size=11)
    ws[f'A{row}'].alignment = center_align
    
    row += 2
    
    # Summary section
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Total Deposits:"
    ws[f'B{row}'] = f"${total_deposits:.2f}"
    ws[f'B{row}'].font = Font(bold=True, color="27ae60")
    row += 1
    
    ws[f'A{row}'] = "Total Debits:"
    ws[f'B{row}'] = f"${total_debits:.2f}"
    ws[f'B{row}'].font = Font(bold=True, color="e74c3c")
    row += 1
    
    ws[f'A{row}'] = "Ending Balance:"
    ws[f'B{row}'] = f"${balance:.2f}"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="2980b9")
    row += 2
    
    # Category breakdown
    if category_totals:
        ws[f'A{row}'] = "Spending by Category"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for cat, amount in sorted(category_totals.items()):
            ws[f'A{row}'] = cat.title() + ":"
            ws[f'B{row}'] = f"${amount:.2f}"
            row += 1
        
        row += 1
    
    # Transactions table headers
    headers = ['Date', 'Description', 'Merchant', 'Category', 'Type', 'Amount', 'Location']
    ws[f'A{row}'] = headers[0]
    ws[f'B{row}'] = headers[1]
    ws[f'C{row}'] = headers[2]
    ws[f'D{row}'] = headers[3]
    ws[f'E{row}'] = headers[4]
    ws[f'F{row}'] = headers[5]
    ws[f'G{row}'] = headers[6]
    
    # Style headers
    for col in range(1, 8):
        cell = ws[get_column_letter(col) + str(row)]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    row += 1
    
    # Add transaction data
    for trans in transactions:
        trans_date = trans[1] if isinstance(trans[1], str) else trans[1].strftime('%Y-%m-%d')
        location = f"{trans[7] or ''}, {trans[8] or ''}".strip(', ')
        
        ws[f'A{row}'] = trans_date
        ws[f'B{row}'] = trans[2]  # description
        ws[f'C{row}'] = trans[3] or ''  # merchant
        ws[f'D{row}'] = trans[4].title()  # category
        ws[f'E{row}'] = trans[6].upper()  # type
        ws[f'F{row}'] = trans[5]  # amount
        ws[f'G{row}'] = location
        
        # Format amount column
        amount_cell = ws[f'F{row}']
        amount_cell.number_format = '$#,##0.00'
        if trans[6] == 'deposit':
            amount_cell.font = Font(bold=True, color="27ae60")
        else:
            amount_cell.font = Font(bold=True, color="e74c3c")
        
        # Apply borders to all cells
        for col in range(1, 8):
            cell = ws[get_column_letter(col) + str(row)]
            cell.border = border
            if col == 5:  # Type column
                cell.alignment = center_align
            elif col == 6:  # Amount column
                cell.alignment = right_align
        
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 35  # Description
    ws.column_dimensions['C'].width = 20  # Merchant
    ws.column_dimensions['D'].width = 15  # Category
    ws.column_dimensions['E'].width = 10  # Type
    ws.column_dimensions['F'].width = 15  # Amount
    ws.column_dimensions['G'].width = 25  # Location
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"transactions_check_{paystub_info[1]}_{paystub_info[2]}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
